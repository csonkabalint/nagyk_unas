import openpyxl
from openpyxl import load_workbook
import re
from util import eval_stock, create_new_unas_full
from util import on_stock


class ItemHasNoSexException(Exception):
    """Raised when item no sex"""
    pass


class ItemHasNoConcentrationException(Exception):
    """Raised when item no concentration"""
    pass


class ItemHasNoSizeException(Exception):
    """Raised when item no size"""
    pass


all_type_tag = ['flacon', 'shower gel', 'set', 'body scrub', 'candle', 'bl',
                'http://www.upcitemdb.com/upc/8028713470035', 'vial', 'woman', 'falcon', 'deo', 'as', 'damaged',
                'stick', 'body lotion', 'mini set', 'bodymilk', 'mini', 'showergel', 'niche', 'sg', 'bodylotion',
                'set 2017', 'set 2018', 'bodymist', 'no celo', 'de', 'spray prob', 'no box', 'mini set 5p', 'flacon as',
                'bcream', '088300170814', 'set 2016', 'body mist', 'roll on', 'bodycream', 'set 2015', 'flaconas',
                'cut code', 'spray head', 'asb flac.', 'asgel', 'asb', 'dushgel', 'soap', 'ab', 'cream', 'hairmist',
                'foam bath', '6 g', 'hair spray', 'stiff', 'roller', 'savinggél', 'giftbox.', 'asb flacon', 'as gel',
                'flaconb.mist', 'flaconbl', 'damamged', 'bl flacon', 'body moussre', 'body cream',
                'set 2022', 'pen', '', 'sg cream', 'exclusive xmas set', '027131450986', '027131669111', '027131669142',
                '027131495284', '027131695660', '027131317791', '027131365471', '027131365488', '027131342267',
                '027131669135', '027131668930', '027131668947', '027131668978', '027131668985', '027131759959',
                '027131317777', '027131499978', '027131674863', '027131674900', '027131674962', '027131582823',
                '027131829973', '027131829997', '027131830016', '027131830054', '027131830115', '027131830160',
                '027131669067', '027131669081', '027131669098', '027131669128', '027131782018', '027131830474',
                '027131830634', '027131830672', '027131489511', '027131835554', '027131696438', '027131674894',
                'perfect peel', '027131844143', '027131752004', '027131830412', '027131830238', '027131830184',
                '027131759423', '027131751922', '027131003755', '027131208709', '027131674870', '027131781738',
                '027131781721', '027131696537', 'cosm.', '027131632849', '10packettes', 'shaving', 'not full',
                'body spray', 'stift', 'foambath', 'hair gel', 'as flacon', 'shaving foam', 'dushgél',
                'dushgél flakon', 'showgel', 'deo roll', 'asb damaged', 'handcream', 'flacon bl', 'flscon',
                'showercream', 'bag', 'de code', 'set 2019', 'new 2018', 'asb fl',
                'https://www.google.hu/search?q=3614270823886&client=firefox-b&source=lnms&tbm=isch&sa=x&ved=0ahukewivn_mxgmtsahvejpokhfyparuq_auicsgc&biw=1535&bih=805#imgrc=xn_8gfq_zaac5m:',
                'damaged deo', '3605530856157', '3147758014198', '3605530362238', '3147753565015', '3605532202945',
                '3147758185010', '3605531662023', '3147758097108', 'body oil', 'bodylorion', 'fl sg', 'hair mist',
                'aftershave', 'set 2021', 'set mini', 'bodylotion flacon', 'wallet',
                'shampoo', 'roll-on', 'shower cream', 'serum', 'cream set', 'man', 'sérium',
                'serum set', 'blflacon', 'damage', 'hand cream', 'tube', 'spray', 'set 2014',
                'https://photos.google.com/share/af1qipmmx33ighe6upipa8y3qhokwny97vlnqzydq6iawliek6m4ligbw46zzyu7qwdehg/photo/af1qipojrt4jfdteo10xq9juxlpl5tsy65k-etoaguad?key=ytu1qwnfnuvmq1jltc1heuppqulsv05tuhowd2rn',
                'cutcode', 'parf deo', 'set 2020', 'bodylo. flacon']

# column indexes
nagyk_refnum_col = 1
nagyk_brand_col = 2
nagyk_title_col = 3
nagyk_package_col = 4
nagyk_sex_col = 5
nagyk_size_col = 6
nagyk_stock_col = 10
nagyk_price_col = 9

unas_refnum_col = 0
unas_nett_col = 2
unas_brutt_col = 3
unas_akc_nett_col = 4
unas_akc_brutt_col = 5

# price conversion rates
disc_rate = 0.95
afa = 1.27

# stock labels
empty_stock_labels = ["On stock 1B", "On stock 1 B", "None"]


def nagyk_nett_unas_brutt(nagyker_nett):
    return nagyker_nett * 1.4605


def discount(prc):
    return prc * disc_rate


def brutt_nett(prc):
    return prc / afa


def calc_prices(nagyker_nett):
    unas_brutt = nagyk_nett_unas_brutt(nagyker_nett)
    unas_nett = brutt_nett(unas_brutt)
    unas_brutt_akc = discount(unas_brutt)
    unas_nett_akc = discount(unas_nett)
    """ rounding
    unas_brutt = round(unas_brutt)
    unas_nett = round(unas_nett)
    unas_brutt_akc = round(unas_brutt_akc)
    unas_nett_akc = round(unas_nett_akc)
    """
    return unas_nett, unas_brutt, unas_nett_akc, unas_brutt_akc


relevant_parfume_types = ["parfume", "cologne", "toalett"]
not_relevant_package_types = ["set", "flacon", "damaged", "special", "body lotion", "not full", "candle", "shower gel",
                              "scrub", "showergel", "sg", "body lotion", "falcon"]


def is_relevant(relevant_parfume_types, not_relevant_package_types, package_type, parfume_type="parfume"):
    if parfume_type is not None:
        parfume_type = parfume_type.lower().strip()
    if package_type is not None:
        package_type = package_type.lower().strip()
    if parfume_type in relevant_parfume_types and package_type is None or not any(
            pac in package_type for pac in not_relevant_package_types):
        return True
    else:
        return False


def correct_reference(reference):
    if reference.startswith("D-G"):
        reference = reference.replace("-", "&")
    return reference



header_labels = 	"Paraméter: Mennyiség||text", "Paraméter: Márka||text",	"Paraméter: Parfümök||text",	"Paraméter: Szállítási idő||text",	"Paraméter: Szállítási költség||text",	"Paraméter: Kiegészítő termékek||text",	"Paraméter: Arukereso.hu Export Kategória||text",	"Alternatív Kategória 1",	"Alternatív Kategória 2",	"Alternatív Kategória 3",	"Alternatív Kategória 4"



# nem mukodik?
def add_unas_row(unas, row, cikkszam, termek_nev, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, akc_kezd, akc_lejar,
                 rakt_kesz, vasar_ha_nincs_rakt=0, valtozat_rakt_kesz=0, alacsony_kesz=None):
    params = locals()
    print(locals())
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + row
        unas.active[pos] = params[idx + 2]


def generate_unas_title(unas_brand, nagyk_brand, title, size, package, sex):
    no_sex = True
    no_concentration = True
    no_size = True

    title_unas = unas_brand + " "

    """
    test_title = remove_substring_concentration(title)
    test_title += get_concentration_from_title(title)
    print(test_title)
    """
    title_removed = remove_brand_name(unas_brand, nagyk_brand, title)
    title_removed = remove_concentration(title_removed)
    title_removed = remove_substring_concentration(title_removed)
    title_unas += title_removed + " "

    sex_extr = translate_sex(sex)
    if sex_extr is not None:
        title_unas += sex_extr
        no_sex = False

    concentration = translate_concentration(title)
    if concentration is not None:
        title_unas += concentration
        no_concentration = False
    else:
        concentration = get_concentration_from_title(title)
        if concentration is not None:
            title_unas += concentration
            no_concentration = False

    size = translate_size(size)
    if size is not None:
        title_unas += size
        no_size = False

    # TODO testert rendesen csekkolni
    if "tester" in title.lower() or package is not None and "tester" in package.lower():
        title_unas.replace("Tester Fragrence", "")
        title_unas += "teszter"

    title_unas = re.sub('\s+', ' ', title_unas).strip()
    return title_unas, no_sex, no_concentration, no_size


def generate_unas_parameters(unas_brand, nagyk_brand, title, size_extr, package, sex):

    title_extr = remove_brand_name(unas_brand, nagyk_brand, title)
    # title_extr = remove_concentration(title_extr)
    title_extr = remove_substring_concentration(title_extr)

    sex_extr = translate_sex(sex)

    concentration_extr = get_concentration_from_title(title)

    size_extr = translate_size(size_extr)

    tester = None

    # TODO testert rendesen csekkolni, kulon fvben
    if "tester" in title.lower() or package is not None and "tester" in package.lower():
        tester = "teszter"

    if title_extr is not None:
        title_extr = re.sub('\s+', ' ', title_extr).strip()
    if sex_extr is not None:
        sex_extr = re.sub('\s+', ' ', sex_extr).strip()
    if concentration_extr is not None:
        concentration_extr = re.sub('\s+', ' ', concentration_extr).strip()
    if tester is not None:
        tester = re.sub('\s+', ' ', tester).strip()

    return title_extr, sex_extr, concentration_extr, size_extr

"""
def generate_unas_title(unas_brand, title, size, concentration, package, sex):
    title_unas = unas_brand + " "
    title_unas += title
    title_unas += sex
    title_unas += concentration
    title_unas += size

    if "tester" in title.lower() or package is not None and "tester" in package.lower():
        title_unas.replace("Tester Fragrence", "")
        title_unas += "teszter"

    title_unas = re.sub('\s+',' ',title_unas).strip()
    return title_unas
"""


def translate_sex(sex):
    if sex is None:
        return None
    sex = sex.lower()
    if sex == "man":
        return "férfi "
    elif sex == "woman":
        return "női "
    elif sex == "unisex":
        return "unisex "


def get_concentration_from_title(title):

    title = title.lower()

    parf = ["edp", "parfum"]
    colo = ["edc", "cologne"]
    toil = ["edt", "toilett"]

    if any(tag in title for tag in parf):
        return 'eau de parfum '
    if any(tag in title for tag in colo):
        return 'eau de cologne '
    if any(tag in title for tag in toil):
        return 'eau de toilette '
    return None


def remove_concentration_from_title(title):
    insensitive_edp = re.compile(re.escape('eau de parfum'), re.IGNORECASE)
    title = insensitive_edp.sub("", title)
    insensitive_edc = re.compile(re.escape('eau de cologne'), re.IGNORECASE)
    title = insensitive_edc.sub("", title)
    insensitive_edt = re.compile(re.escape('eau de toilette'), re.IGNORECASE)
    title = insensitive_edt.sub("", title)
    return title


def translate_concentration(title):
    title = title.lower()
    if "edp" in title:
        return "eau de parfume "
    elif "edc" in title:
        return "eau de cologne "
    elif "edt" in title:
        return "eau de toilette "
    else:
        None


def translate_size(size):
    if size is not None:
        return size + " "
    else:
        return None


def remove_brand_name(unas_brand, nagyk_brand, title):
    title = title.replace(unas_brand.strip(), "")
    title = title.replace(nagyk_brand.strip(), "")
    return title


def remove_substring_concentration(string):
    substrings_to_remove = ["edp", "edc", "edt", "eau de parfume", "eau de cologne", "eau de toilette", "eau de parfum", "parfume", "parfum", "toilette", "cologne"]
    for substring in substrings_to_remove:
        string = re.sub(substring, "", string, flags=re.IGNORECASE)
    return string


replace_concentration_tags = {"edp": "",
                         "edc" : "",
                         "edt" : "",
                         "eau de parfume" : "",
                         "eau de cologne" : "",
                         "eau de toilette" : "",
                         "eau de parfum" : "",
                         "parfume" : "",
                         "parfum" : "",
                         "toilette" : "",
                         "cologne" : "", }


def remove_concentration2(title, rep):
    rep = dict((re.escape(k), v) for k, v in rep.items())
    pattern = re.compile("|".join(rep.keys()), re.IGNORECASE)
    text = pattern.sub(lambda m: rep[re.escape(m.group(0))], title)
    return text


def remove_concentration(title):
    insensitive_edp = re.compile(re.escape('edp'), re.IGNORECASE)
    title = insensitive_edp.sub("", title)
    insensitive_edc = re.compile(re.escape('edc'), re.IGNORECASE)
    title = insensitive_edc.sub("", title)
    insensitive_edt = re.compile(re.escape('edt'), re.IGNORECASE)
    title = insensitive_edt.sub("", title)
    return title


def add_unas_row(unas, row, cikkszam, termek_nev, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, rakt_kesz,
                 akc_kezd=None, akc_lejar=None, vasar_ha_nincs_rakt=0, valtozat_rakt_kesz=0, alacsony_kesz=None):
    str_row = str(row)
    unas.active["A" + str_row] = str(cikkszam)
    unas.active["B" + str_row] = termek_nev
    unas.active["C" + str_row] = netto_ar
    unas.active["D" + str_row] = brutto_ar
    unas.active["E" + str_row] = akc_netto_ar
    unas.active["F" + str_row] = akc_brutto_ar
    unas.active["I" + str_row] = rakt_kesz


def add_unas_row_ext(unas, row, cikkszam, termek_nev, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, ketegoria, rakt_kesz, meret, marka, nem,
                     leiras=None, akc_kezd=None, akc_lejar=None, vasar_ha_nincs_rakt=0, valtozat_rakt_kesz=0, alacsony_kesz=None):
    str_row = str(row)
    unas.active["A" + str_row] = str(cikkszam)
    unas.active["B" + str_row] = termek_nev
    unas.active["C" + str_row] = netto_ar
    unas.active["D" + str_row] = brutto_ar
    unas.active["E" + str_row] = akc_netto_ar
    unas.active["F" + str_row] = akc_brutto_ar
    unas.active["I" + str_row] = ketegoria
    # unas.active["J" + str_row] = leiras
    unas.active["K" + str_row] = rakt_kesz
    unas.active["L" + str_row] = 0
    unas.active["M" + str_row] = 0
    unas.active["O" + str_row] = meret
    unas.active["P" + str_row] = marka
    unas.active["Q" + str_row] = nem
    unas.active["V" + str_row] = "Márkák|" + marka


def add_unas_row_simplified2(unas, unas_old, row, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, rakt_kesz):
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        unas.active[pos] = unas_old.active[pos].value
    str_row = str(row)
    unas.active["C" + str_row] = netto_ar
    unas.active["D" + str_row] = brutto_ar
    unas.active["E" + str_row] = akc_netto_ar
    unas.active["F" + str_row] = akc_brutto_ar
    unas.active["K" + str_row] = rakt_kesz


def unas_row_copy(unas, unas_old, row):
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        unas.active[pos] = unas_old.active[pos].value
    str_row = str(row)


def print_unas_row(unas, row):
    alpha = "ABCDEFGHIJKL"
    to_print = ""
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        to_print += str(unas.active[pos].value)
        to_print += " | "
    print(to_print)


brands = ["Abercrombie & Fitch",
          "Acqua Di Parma",
          "Ajmal",
          "Amouage",
          "Antonio Banderas",
          "Balenciaga",
          "Boucheron",
          "Burberry",
          "Bvlgari",
          "Byredo",
          "Cacharel",
          "Calvin Klein",
          "Carolina Herrera",
          "Cerruti",
          "Chanel",
          "Chloé",
          "Christian Dior",
          "Creed",
          "Davidoff",
          "Diesel",
          "DKNY",
          "Dolce & Gabbana",
          "Dsquared",
          "Elie Saab",
          "Elizabeth Arden",
          "Escada",
          "Estée Lauder",
          "Etat Libre d’Orange",
          "Gianfranco Ferr",
          "Giorgio Armani",
          "Givenchy",
          "Gucci",
          "Guerlain",
          "Guess",
          "Hermes",
          "Hugo Boss",
          "Iceberg",
          "Issey Miyake",
          "Jean Paul Gaultier",
          "Jimmy Choo",
          "Joop",
          "Karl Lagerfeld",
          "Kenzo",
          "Kilian",
          "Lacoste",
          "Lalique",
          "Lancome",
          "Lanvin",
          "Lolita lempicka",
          "Maison Francis Kurkjian Paris",
          "Mancera",
          "Marc Jacobs",
          "Michael Kors",
          "Montale",
          "Mont Blanc",
          "Moschino",
          "Narciso Rodriguez",
          "Nasomatto",
          "Nicolai",
          "Nina Ricci",
          "Nishane",
          "Paco Rabanne",
          "Parfums De Marly",
          "Philipp Plein",
          "Prada",
          "Rasasi",
          "Roberto Cavalli",
          "Roja Parfums",
          "Shiseido",
          "Thierry Mugler",
          "Tiziana Terenzi",
          "Tom Ford",
          "Trussardi",
          "Valentino",
          "Versace",
          "Viktor & Rolf",
          "Xerjoff",
          "Yves S. L."]

nagyk_brands = ["Abercrombie & Fitch",
                "Acqua Di Parma",
                "Ajmal",
                "Amouage",
                "Antonio Banderas",
                "Balenciaga",
                "Boucheron",
                "Burberry",
                "Bvlgari",
                "Byredo",
                "Cacharel",
                "Calvin Klein",
                "Carolina Herrera",
                "Cerruti",
                "Chanel",
                "Chloé",
                "Christian Dior",
                "Creed",
                "Davidoff",
                "Diesel",
                "DKNY",
                "Dolce & Gabbana",
                "Dsquared",
                "Elie Saab",
                "Elizabeth Arden",
                "Escada",
                "Estée Lauder",
                "Etat Libre o'range",
                "Gianfranco Ferre",
                "Giorgio Armani",
                "Givenchy",
                "Gucci",
                "Guerlain",
                "Guess",
                "Hermes",
                "Hugo Boss",
                "Iceberg",
                "Issey myake",
                "JPG",
                "Jimmy Choo",
                "Joop",
                "Lagerfeld",
                "Kenzo",
                "Kilian",
                "Lacoste",
                "Lalique",
                "Lancome",
                "Lanvin",
                "Lempicka",
                "Maison Francis Kurkjian Paris",
                "Mancera",
                "Marc Jacobs",
                "Michael Kors",
                "Montale",
                "Mont Blanc",
                "Moschino",
                "Narciso Rodrigez",
                "Nasomatto",
                "Nicolai",
                "Nina Ricci",
                "Nishane",
                "Paco Rabanne",
                "Marly",
                "Philipp Plein",
                "Prada",
                "Rasasi",
                "Roberto Cavalli",
                "Roja",
                "Shiseido",
                "Thierry Mugler",
                "Tiziana Terenzi",
                "Tom Ford",
                "Trussardi",
                "Valentino",
                "Versace",
                "Viktor & Rolf",
                "Xerjoff",
                "Yves S. L."]


def check_refnum(refnum, reference):
    # print(refnum)
    if "nincs" in str(refnum): #not isinstance(refnum, int) and not isinstance(refnum, float) and "nincs" in refnum.strip():
        return reference.replace("&", "-")
    else:
        return refnum


# TODO arkerekites
if __name__ == '__main__':
    print("nagyk_unas start\n")

    all_package_type = []

    wb_unas_new = create_new_unas_full()
    wb_unas_missing_attr = create_new_unas_full()

    # wb_unas = load_workbook(filename="UNAS-EXPORT.xlsx", data_only=True)
    wb_nagyk = load_workbook(filename="data/NAGYKER_ARLISTA.xlsx", data_only=True)

    count_instances = 0
    count_instances_missing = 0

    for idx_brand, nagyk_brand in enumerate(nagyk_brands):
        print(nagyk_brand)
        print('##################################')
        for idx, row in enumerate(wb_nagyk.active.iter_rows(min_col=2, max_col=3, min_row=1)):  # , max_row=70
            idx_one_indexed = idx + 1
            if str(row[1].value).lower().strip() == nagyk_brand.lower():
                row_nagyk = wb_nagyk.active[idx_one_indexed]
                price = row_nagyk[nagyk_price_col].value
                stock = row_nagyk[nagyk_stock_col].value
                package = row_nagyk[nagyk_package_col].value
                if package is not None and package.lower().strip() not in all_package_type:
                    all_package_type.append(package.lower().strip())

                if price is not None and not isinstance(price, str) and price > 13000 and on_stock(stock,
                                                                                                   empty_stock_labels) and is_relevant(
                        relevant_parfume_types, all_package_type, package):
                    ean_num = row_nagyk[nagyk_refnum_col].value
                    reference = row_nagyk[0].value
                    ean_num = check_refnum(ean_num, reference)
                    unas_brand = brands[idx_brand]
                    nagyk_title = row_nagyk[nagyk_title_col].value
                    size = row_nagyk[nagyk_size_col].value
                    sex = row_nagyk[nagyk_sex_col].value
                    stock_eval = eval_stock(stock, empty_stock_labels)

                    title, sex, concentration, size = generate_unas_parameters(unas_brand, nagyk_brand, nagyk_title, size, package, sex)

                    netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar = calc_prices(float(price))

                    if sex is not None and concentration is not None and size is not None:

                        unas_title = unas_brand + " " + title + " " + sex + " " + concentration + " " + size
                        unas_category = "Parfümök|" + sex.capitalize() + " parfümök|" + concentration
                        # print(unas_category)

                        count_instances += 1
                        add_unas_row_ext(wb_unas_new, count_instances + 1, ean_num, unas_title, netto_ar, brutto_ar,
                                 akc_netto_ar, akc_brutto_ar, unas_category, stock_eval, size, unas_brand, sex.capitalize())


                    else:
                        unas_title = unas_brand + " " + title + " " + str(sex or '') + " " + str(concentration or '') + " " + str(size or '')
                        if size is None:
                            unas_title += " NINCS MÉRET"
                        if concentration is None:
                            unas_title += " NINCS KONCENTRÁCIÓ"
                        if sex is None:
                            unas_title += " NINCS NEM"
                        count_instances_missing += 1
                        add_unas_row(wb_unas_missing_attr, count_instances_missing + 1, ean_num, unas_title, netto_ar, brutto_ar,
                                     akc_netto_ar, akc_brutto_ar, stock_eval)
                    #print(unas_title)
                    # print(row_nagyk[2].value, row_nagyk[3].value)
                #else:
                    #print_unas_row(wb_nagyk, idx_one_indexed)
        # print("--------------------------")
    print(count_instances)
    print(count_instances_missing)
    wb_unas_new.save("nagyk_levalogatas.xlsx")
    wb_unas_missing_attr.save("wb_unas_missing.xlsx")
    print(all_package_type)
