import openpyxl
from openpyxl import load_workbook


brands = ["Abercrombie & Fitch",
"Acqua Di Parma",
"Ajmal",
"Amouage",
"Antonio Banderas",
"Balenciaga",
"Boucheron",
"Burberry",
"Bvlgari",
"Byredo",
"Cacharel",
"Calvin Klein",
"Carolina Herrera",
"Cerruti",
"Chanel",
"Chloé",
"Christian Dior",
"Creed",
"Davidoff",
"Diesel",
"DKNY",
"Dolce & Gabbana",
"Dsquared",
"Elie Saab",
"Elizabeth Arden",
"Escada",
"Estée Lauder",
"Etat Libre d’Orange",
"Gianfranco Ferr",
"Giorgio Armani",
"Givenchy",
"Gucci",
"Guerlain",
"Guess",
"Hermes",
"Hugo Boss",
"Iceberg",
"Issey Miyake",
"Jean Paul Gaultier",
"Jimmy Choo",
"Joop",
"Karl Lagerfeld",
"Kenzo",
"Kilian",
"Lacoste",
"Lalique",
"Lancome",
"Lanvin",
"Lolita lempicka",
"Maison Francis Kurkjian Paris",
"Mancera",
"Marc Jacobs",
"Michael Kors",
"Montale",
"Mont Blanc",
"Moschino",
"Narciso Rodriguez",
"Nasomatto",
"Nicolai",
"Nina Ricci",
"Nishane",
"Paco Rabanne",
"Parfums De Marly",
"Philipp Plein",
"Prada",
"Rasasi",
"Roberto Cavalli",
"Roja Parfums",
"Shiseido",
"Thierry Mugler",
"Tiziana Terenzi",
"Tom Ford",
"Trussardi",
"Valentino",
"Versace",
"Viktor & Rolf",
"Xerjoff",
"Yves S. L."]


def check_brand(brands, nk_wb):
    found = False
    for brand in brands:
        print(brand)
        for row in enumerate(nk_wb.iter_rows(min_col=3, max_col=3, min_row=1)):
            if str(row[1][0].value).lower().strip() == brand.lower().strip():
                print("found: " + row[1][0].value)
                found = True
                break
        if not found:
            print("not found:" + brand)
        found = False
        print("_____________________________")


if __name__ == '__main__':
    print("nagyk_check_brands start\n")

    wb_nagyk = load_workbook(filename="data/NAGYKER_ARLISTA.xlsx", data_only=True)
    check_brand(brands, wb_nagyk.active)
