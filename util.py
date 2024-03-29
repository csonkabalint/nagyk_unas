from datetime import date
from string import ascii_uppercase
import openpyxl
import tkinter as tk
from tkinter import filedialog



class ItemNotFoundException(Exception):
    """Raised when item not found"""
    pass



def get_row_index(cell_value_tofind, work_book):
    for idx, row in enumerate(work_book.iter_rows(min_col=1, max_col=2, min_row=1)):
        for cell in row:
            result =str(cell.value).split(".", 1)[0]  ## kezeli a tizedejegyeket át kéne máshova szervezni!!!
            if str(result) == cell_value_tofind:
                print("found")
                pr_row = ""
                for c in row:
                    pr_row += str(c.value)
                    pr_row += ", "
                print(pr_row, idx)
                return idx + 1
    raise ItemNotFoundException


def print_line(print_on, *text):
    if print_on:
        print(text)


def new_unas_row(unas, row_idx, *args):
    row_idx_str = str(row_idx)
    for idx, arg in enumerate(args):
        pos = ascii_uppercase[idx] + row_idx_str
        unas.active[pos] = arg


def eval_stock(stock_label, empty_labels):
    for label in empty_labels:
        if label == stock_label:
            return 0
    return 1000


def on_stock(stock_label, empty_labels):
    for label in empty_labels:
        if label == stock_label:
            return False
    return True

def create_new_unas_full():
    unas = openpyxl.Workbook()
    header_labels = ["Cikkszám", "Termék Név", "Nettó Ár", "Bruttó Ár", "Akciós Nettó Ár", "Akciós Bruttó Ár",
                     "Akció Kezdet", "Akció Lejárat", "Kategória", "Rövid Leírás", "Raktárkészlet", "Vásárolható, ha nincs Raktáron",
                     "Változatokhoz Raktárkészlet", "Alacsony készlet", "Paraméter: Mennyiség||text", "Paraméter: Márka||text",
                     "Paraméter: Parfümök||text",	"Paraméter: Szállítási idő||text",	"Paraméter: Szállítási költség||text",
                     "Paraméter: Kiegészítő termékek||text",	"Paraméter: Arukereso.hu Export Kategória||text",	"Alternatív Kategória 1",
                     "Alternatív Kategória 2",	"Alternatív Kategória 3",	"Alternatív Kategória 4"]
    # TODO "range char"-ra lecserelni ezt a megoldast
    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXY"
    for idx, label in enumerate(header_labels):
        pos = alpha[idx] + "1"
        unas.active[pos] = header_labels[idx]
    return unas

def create_new_unas_partial():
    unas = openpyxl.Workbook()
    header_labels = ["Cikkszám", "Termék Név", "Nettó Ár", "Bruttó Ár", "Akciós Nettó Ár", "Akciós Bruttó Ár", "Akció Kezdet", "Akció Lejárat", "Raktárkészlet", "Vásárolható, ha nincs Raktáron", "Változatokhoz Raktárkészlet", "Alacsony készlet"]
    # TODO "range char"-ra lecserelni ezt a megoldast
    alpha = "ABCDEFGHIJKL"
    for idx, label in enumerate(header_labels):
        pos = alpha[idx] + "1"
        unas.active[pos] = header_labels[idx]
    return unas


def create_new_unas_partial_stock_update():
    unas = openpyxl.Workbook()
    header_labels = ["Cikkszám", "Termék Név", "Nettó Ár", "Bruttó Ár", "Akciós Nettó Ár", "Akciós Bruttó Ár", "Akció Kezdet", "Akció Lejárat", "Kategória", "Raktárkészlet", "Vásárolható, ha nincs Raktáron", "Változatokhoz Raktárkészlet", "Alacsony készlet"]
    # TODO "range char"-ra lecserelni ezt a megoldast
    alpha = "ABCDEFGHIJKLM"
    for idx, label in enumerate(header_labels):
        pos = alpha[idx] + "1"
        unas.active[pos] = header_labels[idx]
    return unas

def get_file_path_dialog(header_text, default_dir):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(initialdir=default_dir,title=header_text)
    return file_path


def filename_with_date(filename, file_extension):
    date_string = str(date.today()).replace("-", "_")
    return filename + "_" + date_string + "." + file_extension