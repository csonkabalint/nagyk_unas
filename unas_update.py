import os
from openpyxl import load_workbook
import util
from util import print_line, get_row_index, ItemNotFoundException, create_new_unas_partial
from util import eval_stock


print_on = False  # True


# column indexes
nagyk_stock_col = 10
nagyk_price_col = 9

unas_refnum_col = 0
unas_nett_col = 2
unas_brutt_col = 3
unas_akc_nett_col = 4
unas_akc_brutt_col = 5

# price conversion rates
disc_rate = 0.95
afa = 1.27

# stock labels
empty_stock_labels = ["On stock 1B", "On stock 1 B", "None"]


def nagyk_nett_unas_brutt(nagyker_nett):
    return nagyker_nett * 1.4605


def discount(prc):
    return prc * disc_rate


def brutt_nett(prc):
    return prc / afa


def calc_prices(nagyker_nett):
    unas_brutt = nagyk_nett_unas_brutt(nagyker_nett)
    unas_nett = brutt_nett(unas_brutt)
    unas_brutt_akc = discount(unas_brutt)
    unas_nett_akc = discount(unas_nett)
    """ TODO? rounding?
    unas_brutt = round(unas_brutt)
    unas_nett = round(unas_nett)
    unas_brutt_akc = round(unas_brutt_akc)
    unas_nett_akc = round(unas_nett_akc)
    """
    return unas_nett, unas_brutt, unas_nett_akc, unas_brutt_akc


def correct_reference(reference):
    if reference.startswith("D-G"):
        reference = reference.replace("-", "&")
    return reference






def add_unas_row_simplified2(unas, unas_old, row, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, rakt_kesz):
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        unas.active[pos] = unas_old.active[pos].value
    str_row = str(row)
    unas.active["C" + str_row] = netto_ar
    unas.active["D" + str_row] = brutto_ar
    unas.active["E" + str_row] = akc_netto_ar
    unas.active["F" + str_row] = akc_brutto_ar
    unas.active["I" + str_row] = rakt_kesz


def unas_row_copy(unas, unas_old, row):
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        unas.active[pos] = unas_old.active[pos].value
    str_row = str(row)


def unas_row_copy_stock(unas, unas_old, row, stock):
    alpha = "ABCDEFGHIJKL"
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        unas.active[pos] = unas_old.active[pos].value
        unas.active["I" + str(row)] = stock
    str_row = str(row)


def print_unas_row(unas, row):
    alpha = "ABCDEFGHIJKL"
    to_print = ""
    for idx, cell_col in enumerate(alpha):
        pos = cell_col + str(row)
        to_print += str(unas.active[pos].value)
        to_print += " | "
    print(to_print)


if __name__ == '__main__':
    print("nagyk_unas start\n")

    wb_unas_not_found = create_new_unas_partial()
    wb_unas_new = create_new_unas_partial()

    wb_unas = load_workbook(filename=util.get_file_path_dialog("open unas", os.getcwd()), data_only=True)
    wb_nagyk = load_workbook(filename=util.get_file_path_dialog("open nagyk", os.getcwd()), data_only=True)

    for idx, row in enumerate(wb_unas.active.iter_rows(min_col=1, max_col=1, min_row=2)):
        to_find = str(row[unas_refnum_col].value)
        to_find = correct_reference(to_find)
        print_line(print_on, "to find: " + to_find)

        try:
            row_idx = get_row_index(to_find, wb_nagyk.active)
            row_nagyk = wb_nagyk.active[row_idx]
            price_nagyk = str(row_nagyk[nagyk_price_col].value)
            stock_nagyk = str(row_nagyk[nagyk_stock_col].value)
            netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar = calc_prices(float(price_nagyk))
            stock_unas = eval_stock(stock_nagyk, empty_stock_labels)
            add_unas_row_simplified2(wb_unas_new, wb_unas, idx + 1 + 1, netto_ar, brutto_ar, akc_netto_ar, akc_brutto_ar, stock_unas)
            print_unas_row(wb_unas_new, idx + 2)

        except ItemNotFoundException:
            print(str(ItemNotFoundException))
            # TODO kezelni a meg nem talalt rekordokat
            unas_row_copy_stock(wb_unas_new, wb_unas, idx + 1 + 1, 0)
            unas_row_copy(wb_unas_not_found, wb_unas, idx + 1 + 1)
        print_line(print_on, "--------------------------")
    wb_unas_new.save("unas_" + date.today().replace("-", "_") + ".xlsx")
    wb_unas_not_found.save("wb_unas_not_found_" + date.today().replace("-", "_") + ".xlsx")
